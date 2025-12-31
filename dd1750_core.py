import math, os, re, csv
from collections import OrderedDict

import fitz  # PyMuPDF
from PIL import Image
import pytesseract
import openpyxl
import yaml

from pypdf import PdfReader, PdfWriter, PageObject
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch

# Phrases to ignore when scanning BOM text
BADWORDS = [
    "END ITEM","PUB","PAGE","DATE","IMAGE","LIN","DESC","ACOEI",
    "COMPONENT OF END","SLOC","UIC","NIIN","AUTH","QTYOH",
    "COEI","BOM","LOCATION","LOCID","WTY","ARC","CIIC","UI","SCMC",
    "COMPONENT LISTING","HAND RECEIPT","BASIC ISSUE ITEMS","COMPONENTS OF END ITEM"
]

def load_cfg(path: str) -> dict:
    with open(path, "r") as f:
        return yaml.safe_load(f)

def clean_mat(token: str) -> str:
    token = str(token).split("-C")[0]
    return re.sub(r"[^0-9A-Za-z/\-]", "", token)

def group_words_to_lines(words, y_tol=3.0):
    lines = []
    for w in sorted(words, key=lambda w: (round((w[1]+w[3])/2.0, 1), w[0])):
        x0,y0,x1,y1,t,*_ = w
        y = (y0+y1)/2.0
        if not lines or abs(lines[-1]["y"]-y) > y_tol:
            lines.append({"y": y, "w": [(x0, t)]})
        else:
            lines[-1]["w"].append((x0, t))
    return lines

def _extract_pdf_text_rows(pdf_path: str):
    """
    Generic column-based parser. Works for some BOMs.
    (This is the ONE real function; the public names below are aliases.)
    """
    doc = fitz.open(pdf_path)
    items = []
    for page in doc:
        words = page.get_text("words")
        if not words:
            continue

        W = page.rect.width
        cols = {
            "MAT": (0.06 * W, 0.32 * W),
            "DESC": (0.33 * W, 0.82 * W),
            "QTY": (0.83 * W, 0.98 * W),
        }

        lines = group_words_to_lines(words, y_tol=3.0)
        for L in lines:
            toks = sorted(L["w"], key=lambda z: z[0])
            mat = " ".join([t for x, t in toks if cols["MAT"][0] <= x <= cols["MAT"][1]]).strip()
            desc = " ".join([t for x, t in toks if cols["DESC"][0] <= x <= cols["DESC"][1]]).strip()
            qtys = " ".join([t for x, t in toks if cols["QTY"][0] <= x <= cols["QTY"][1]]).strip()

            if not mat or not desc or not qtys:
                continue

            up = (mat + " " + desc).upper()
            if any(b in up for b in BADWORDS):
                continue

            mat_tok = clean_mat(mat.split()[0])
            m = re.findall(r"\d+", qtys)
            if not m:
                continue
            qty = int(m[-1])
            if qty <= 0:
                continue

            items.append({"mat": mat_tok, "desc": re.sub(r"\s{2,}", " ", desc).strip(), "qty": qty})

    doc.close()
    return items


# ✅ BOTH PUBLIC NAMES EXIST FOREVER (no more NameError)
def extract_pdf_text_rows(pdf_path: str):
    return _extract_pdf_text_rows(pdf_path)

def extract_pdf_text_rows(pdf_path: str):
    return _extract_pdf_text_rows(pdf_path)

# Backwards compatible alias (older code used this name)
extract_pdf_text_rows = extract_pdf_text_rows

def extract_bom_tm_listing(pdf_path: str, qty_max_reasonable: int = 999, page_start: int = 0):
    """Robust parser for TM 'Component Listing / Hand Receipt' PDFs (B49-style)."""
    doc = fitz.open(pdf_path)
    items = []

    mat_digits_re = re.compile(r"^\d{7,12}$")
    mat_alnum_re  = re.compile(r"^[A-Z0-9]{6,}$")
    reject_tok_re = re.compile(r"^(C_|COEI|LV|WTY|ARC|CIIC|SCMC|UI|AUTH|QTY|OH|QTYOH)", re.I)
    qty_re = re.compile(r"(\d+)\s*$")

    def is_good_desc(s: str) -> bool:
        up = s.upper()
        if any(b in up for b in BADWORDS):
            return False
        if "C_" in s or "~" in s:
            return False
        letters = sum(ch.isalpha() for ch in s)
        if letters < 4:
            return False
        if len(s.split()) < 2 and "," not in s:
            return False
        return True

    for page in doc[page_start:]:
        text = page.get_text("text") or ""
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
        i = 0
        while i < len(lines):
            ln = lines[i]
            tok = ln.split()[0] if ln else ""
            if not ln or reject_tok_re.search(tok):
                i += 1
                continue

            token_only = (len(ln.split()) == 1)
            if token_only and (mat_digits_re.match(tok) or mat_alnum_re.match(tok)):
                mat = clean_mat(tok)
                desc = None
                qty = None

                j = i + 1
                while j < min(i + 14, len(lines)):
                    cand = lines[j]
                    if cand.startswith("C_") or "~" in cand:
                        j += 1
                        continue
                    if len(cand) == 1 and cand.isalpha():
                        j += 1
                        continue

                    if desc is None and is_good_desc(cand):
                        desc = cand

                    if qty is None:
                        m = qty_re.search(cand)
                        if m and (" EA " in f" {cand} " or " KT " in f" {cand} "):
                            q = int(m.group(1))
                            if 0 < q <= qty_max_reasonable:
                                qty = q

                    if desc is not None and qty is not None:
                        break
                    j += 1

                if desc and qty:
                    items.append({"mat": mat, "desc": desc, "qty": qty})
                    i = j + 1
                else:
                    i += 1
            else:
                i += 1

    doc.close()
    return items

def ocr_page_items(page, dpi=250, qty_max_reasonable: int = 999):
    pix = page.get_pixmap(dpi=dpi)
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    text = pytesseract.image_to_string(img, config="--psm 6")
    items = []
    qty_re = re.compile(r"(\d+)\s*$")

    for raw in text.splitlines():
        line = raw.strip()
        if not line or len(line) < 6:
            continue
        m = qty_re.search(line)
        if not m:
            continue
        qty = int(m.group(1))
        if qty <= 0 or qty > qty_max_reasonable:
            continue

        first = line.split()[0]
        if not re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9/\-]{2,}", first):
            continue

        desc = re.sub(r"\s+\d+\s*$", "", line[len(first):]).strip(" ,;:-")
        up = (first + " " + desc).upper()
        if any(b in up for b in BADWORDS):
            continue
        if not desc or "~" in desc or "C_" in desc:
            continue

        items.append({"mat": clean_mat(first), "desc": re.sub(r"\s{2,}", " ", desc).strip(), "qty": qty})
    return items

def extract_pdf_ocr_rows(pdf_path: str, dpi=250, page_start=0, page_end=None, qty_max_reasonable: int = 999):
    doc = fitz.open(pdf_path)
    items = []
    end = page_end if page_end is not None else len(doc)
    for i in range(page_start, min(end, len(doc))):
        items.extend(ocr_page_items(doc[i], dpi=dpi, qty_max_reasonable=qty_max_reasonable))
    doc.close()
    return items

def extract_excel_rows(excel_path: str, sheet=None, col_desc="Description", col_mat="Material", col_qty="OH QTY"):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet] if sheet else wb.active

    headers = {}
    for c in ws[1]:
        if c.value:
            headers[str(c.value).strip().lower()] = c.column

    def idx(name): return headers.get(name.lower())

    i_desc, i_mat, i_qty = idx(col_desc), idx(col_mat), idx(col_qty)
    if not (i_desc and i_mat and i_qty):
        raise ValueError(f"Missing columns. Found: {list(headers.keys())}")

    items = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        desc, mat, qty = r[i_desc-1], r[i_mat-1], r[i_qty-1]
        if desc is None or mat is None or qty is None:
            continue
        try:
            qty = int(qty)
        except Exception:
            continue
        if qty <= 0:
            continue

        items.append({"mat": clean_mat(mat), "desc": str(desc).strip(), "qty": qty})
    return items

def aggregate(items):
    agg = OrderedDict()
    for it in items:
        key = (it["mat"], it["desc"])
        agg[key] = agg.get(key, 0) + int(it["qty"])
    return [{"mat": k[0], "desc": k[1], "qty": v} for k,v in agg.items()]

def write_audit(items, out_csv_path, qty_max):
    with open(out_csv_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["line","mat","desc","qty","flag"])
        for i,it in enumerate(items, start=1):
            flag = "SUSPICIOUS_QTY" if it["qty"] > qty_max else ""
            w.writerow([i, it["mat"], it["desc"], it["qty"], flag])

def draw_overlay(overlay_pdf, items, cfg, label="NSN"):
    L = cfg["layout"]
    ipp = int(cfg["items_per_page"])
    unit = cfg["unit_of_issue"]
    spares = cfg["running_spares"]

    c = canvas.Canvas(overlay_pdf, pagesize=letter)

    x_a = L["x_a_center"] * inch
    x_b = L["x_b"] * inch
    x_c = L["x_c_center"] * inch
    x_d = L["x_d_center"] * inch
    x_e = L["x_e"] * inch
    x_f = L["x_f"] * inch

    top = L["top"] * inch
    bottom = L["bottom"] * inch
    y_shift = L["y_block_shift"] * inch

    row_h = (top - bottom) / ipp
    total_pages = max(1, math.ceil(len(items) / ipp))

    for _p in range(total_pages):
        start, end = _p * ipp, min((_p + 1) * ipp, len(items))
        y = top - row_h + 2 - y_shift

        for idx in range(start, end):
            it = items[idx]
            line_no = idx + 1

            c.setFont("Helvetica", 9)
            num = str(line_no)
            c.drawString(x_a - c.stringWidth(num, "Helvetica", 9)/2.0, y, num)

            c.setFont("Helvetica", 9)
            c.drawString(x_b, y, it["desc"][:90])
            c.setFont("Helvetica", 8)
            c.drawString(x_b, y - 12, f"{label}: {it['mat']}")

            c.setFont("Helvetica", 9)
            c.drawString(x_c - c.stringWidth(unit, "Helvetica", 9)/2.0, y, unit)

            q = str(it["qty"])
            c.drawString(x_d - c.stringWidth(q, "Helvetica", 9)/2.0, y, q)
            c.drawString(x_e, y, spares)
            c.drawString(x_f, y, q)

            y -= row_h

        c.showPage()
    c.save()

def merge_template(template_pdf, overlay_pdf, out_pdf):
    reader_template = PdfReader(template_pdf)
    reader_overlay = PdfReader(overlay_pdf)
    writer = PdfWriter()
    base = reader_template.pages[0]

    for ov in reader_overlay.pages:
        merged = PageObject.create_blank_page(width=base.mediabox.width, height=base.mediabox.height)
        merged.merge_page(base)
        merged.merge_page(ov)
        writer.add_page(merged)

    with open(out_pdf, "wb") as f:
        writer.write(f)

def generate_dd1750_from_pdf(
    bom_pdf,
    template_pdf,
    cfg,
    out_pdf,
    out_audit_csv,
    force_ocr=False,
    ocr_dpi=250,
    page_start=0,
    label="NSN"
):
    qty_max_reasonable = int(cfg.get("qty_max_reasonable", 999))

    # 1) Try generic text-based parser
    items = [] if force_ocr else extract_pdf_text_rows(bom_pdf)

    # 2) If too few items, try TM listing parser (B49-style)
    if not force_ocr and len(items) < 10:
        tm_items = extract_bom_tm_listing(bom_pdf, qty_max_reasonable=qty_max_reasonable, page_start=int(page_start))
        if len(tm_items) > len(items):
            items = tm_items

    # 3) OCR only if forced OR nothing found
    if force_ocr or len(items) == 0:
        items = extract_pdf_ocr_rows(bom_pdf, dpi=int(ocr_dpi), page_start=int(page_start), qty_max_reasonable=qty_max_reasonable)

    items = aggregate(items)
    write_audit(items, out_audit_csv, qty_max_reasonable)

    overlay_tmp = os.path.splitext(out_pdf)[0] + "_OVERLAY.pdf"
    draw_overlay(overlay_tmp, items, cfg, label=label)
    merge_template(template_pdf, overlay_tmp, out_pdf)

    return items

def generate_dd1750_from_excel(
    excel_path,
    template_pdf,
    cfg,
    out_pdf,
    out_audit_csv,
    sheet=None,
    col_desc="Description",
    col_mat="Material",
    col_qty="OH QTY",
    label="NSN"
):
    qty_max_reasonable = int(cfg.get("qty_max_reasonable", 999))
    items = extract_excel_rows(excel_path, sheet=sheet, col_desc=col_desc, col_mat=col_mat, col_qty=col_qty)
    items = aggregate(items)
    write_audit(items, out_audit_csv, qty_max_reasonable)

    overlay_tmp = os.path.splitext(out_pdf)[0] + "_OVERLAY.pdf"
    draw_overlay(overlay_tmp, items, cfg, label=label)
    merge_template(template_pdf, overlay_tmp, out_pdf)

    return items

def merge_dd1750_pdfs(pdf_paths, out_pdf, keep_all_pages=False):
    writer = PdfWriter()
    for p in pdf_paths:
        r = PdfReader(p)
        if len(r.pages) == 0:
            continue
        if keep_all_pages:
            for pg in r.pages:
                writer.add_page(pg)
        else:
            writer.add_page(r.pages[0])
    with open(out_pdf, "wb") as f:
        writer.write(f)
